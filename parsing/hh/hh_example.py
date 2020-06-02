import requests
from bs4 import BeautifulSoup
import openpyxl
import csv
import re
import datetime
import os.path


headers = {'accept': '*/*',
           'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'}

MIN_AGE = 30
MAX_AGE = 45
FILE_NAME = "hh_electrician"


def get_date():
    """ Возвращает текущую дату
    """
    now = datetime.datetime.now()
    return str(now.year) + '.' + str(now.month) + '.' + str(now.day)


def get_html(url):
    """ Возвращиет html-код страницы.
    """
    session = requests.Session()
    request = session.get(url, headers=headers)
    return request.text


def get_total_pages(html):
    """ Возвращает колличество найденных страниц
    """
    soup = BeautifulSoup(html, 'lxml')
    pages = soup.find('div', class_='bloko-gap bloko-gap_top')\
                .find_all('a', class_='bloko-button HH-Pager-Control')[-1].get('href')
    total_pages = pages.split('=')[-1]

    return int(total_pages)


def write_csv(data):
    """ Запись полученных данных в файл
    """
    with open(FILE_NAME+'.csv', 'a', encoding='utf8', newline='') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow((data['title'], data['age'], data['pay'], data['experience'], data['url']))


def write_xlsx(data):
    """ Запись полученных данных в файл excel
    """
    if not check_xls_id(get_id(data['url'])):
        if os.path.exists(FILE_NAME+'.xlsx'):
            wb = openpyxl.load_workbook(FILE_NAME+'.xlsx')
            sheet_exists = False
            for name_sheet in wb.sheetnames:
                if name_sheet == get_date():
                    sheet_exists = True
            if sheet_exists:
                ws = wb[get_date()]
                ws_id = wb['ID']
                ws_id.column_dimensions['A'].width = 40
            else:
                ws = wb.create_sheet(get_date())
                ws_id = wb['ID']
                ws.append(('ФИО', 'Возраст', 'Зарплата', 'Стаж', 'Ссылка'))
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 100
                ws['A1'].style = 'Accent1'
                ws['B1'].style = 'Accent1'
                ws['C1'].style = 'Accent1'
                ws['D1'].style = 'Accent1'
                ws['E1'].style = 'Accent1'
        else:
            wb = openpyxl.Workbook()
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                wb.remove(sheet)
            ws_id = wb.create_sheet('ID')
            ws_id.column_dimensions['A'].width = 40
            ws = wb.create_sheet(get_date())
            ws.append(('ФИО', 'Возраст', 'Зарплата', 'Стаж', 'Ссылка'))
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 100
            ws['A1'].style = 'Accent1'
            ws['B1'].style = 'Accent1'
            ws['C1'].style = 'Accent1'
            ws['D1'].style = 'Accent1'
            ws['E1'].style = 'Accent1'
            wb.active = ws
            wb.save(FILE_NAME+'.xlsx')

        ws_id.append((get_id(data['url']), 'id'))
        ws.append((data['title'], data['age'], data['pay'], data['experience'], data['url']))
        wb.active = ws
        wb.save(FILE_NAME+'.xlsx')
        wb.close


def get_id(data):
    """ Получение ID резюме по ссылке
    """
    id_resume = str(data).split('/')[4].split('?')[0]
    return id_resume


def check_xls_id(id_resume):
    """ Проверка отсутствия резюме в базе
    """
    if os.path.exists(FILE_NAME+'.xlsx'):
        wb = openpyxl.load_workbook(FILE_NAME+'.xlsx')
        ws_id = wb['ID']

        for i in range(1, ws_id.max_row + 1):
            if id_resume == ws_id.cell(row=i, column=1).value:
                return True
    return False


def get_page_data(html):
    """ Получение данных со страницы
    """
    soup = BeautifulSoup(html, 'lxml')
    ads = soup.find_all('div', class_='resume-search-item')

    for ad in ads:
        try:
            title = ad.find('div', class_='resume-search-item__header').find('a').text.strip()
        except:
            title = ''
        try:
            url = 'https://barnaul.hh.ru' + \
                ad.find('div', class_='resume-search-item__header').find('a').get('href')
        except:
            url = ''
        try:
            age = ad.find('div', class_='resume-search-item__header')\
                    .find('div', class_='resume-search-item__fullname')\
                    .find('span').text.strip().replace(u'\xa0', u' ').split()[0].strip()
        except:
            age = '1'
        try:
            pay = ad.find('div', class_='resume-search-item__header')\
                    .find('div', class_='resume-search-item__compensation').text.strip().replace(u'\xa0', u' ')
        except:
            pay = ''
        try:
            experience = ad.find(
                'div', class_='resume-search-item__description-content').text.strip().replace(u'\xa0', u' ')
        except:
            experience = ''

        # фильтр для возраста
        if int(age) < MIN_AGE or int(age) > MAX_AGE:
            continue
        # фильтр для слов поиска
        elif re.search(r'\bэлектрик\b', title.lower()) or re.search(r'\электромонтер\b', title.lower()):
            data = {'title': title,
                    'url': url,
                    'age': age,
                    'pay': pay,
                    'experience': experience}
            write_csv(data)
            write_xlsx(data)
        else:
            continue


def main():
    url = 'https://barnaul.hh.ru/search/resume?&area=11&exp_period=all_time&logic=normal&order_by=relevance&pos=full_text&text=%D0%AD%D0%BB%D0%B5%D0%BA%D1%82%D1%80%D0%BE%D0%BC%D0%BE%D0%BD%D1%82%D0%B5%D1%80&page=0'
    base_url = 'https://barnaul.hh.ru/search/resume?&area=11&exp_period=all_time&logic=normal&order_by=relevance&pos=full_text&text=%D0%AD%D0%BB%D0%B5%D0%BA%D1%82%D1%80%D0%BE%D0%BC%D0%BE%D0%BD%D1%82%D0%B5%D1%80&page='

    total_pages = get_total_pages(get_html(url))
    for i in range(0, total_pages):
        url_gen = base_url + str(i)
        html = get_html(url_gen)
        get_page_data(html)
        print('Выполнено: ' + str(i) + ' из ' + str(total_pages))


if __name__ == '__main__':
    main()
